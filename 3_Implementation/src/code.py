import pandas as pd
from openpyxl import load_workbook
import openpyxl
# dataframe is created
df_total = pd.DataFrame()
# all the sheets taken as a list
sheets = ['Sheet1', 'Sheet2','Sheet3','Sheet4','Sheet5']
# from user year is taken
a = int(input("Enter year: "))
# for loop for reading all the data in datasheets and searching for the data we need
for sheet in sheets:
# loop through sheets inside an Excel file
 d = pd.read_excel('New1.xlsx', sheet_name=sheet)
 dataframe = pd.DataFrame(d[d['year'] == a], columns=d.columns)
# here we are joining all the data of that particular person
 df_total = df_total.join(dataframe, how='outer',lsuffix=' ', rsuffix=' ')
# taking the path from the excel sheet
path = "New1.xlsx"
book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book
# searching for the master sheet in all sheets
if 'MasterSheet' in book.sheetnames:
   ref = book['MasterSheet']
   # Removed the previous data in the sheet
   book.remove(ref)
   # write the data into the master sheet
df_total.to_excel(writer, sheet_name='MasterSheet')
# save the data in the master sheet
writer.save()
writer.close()
if 'MasterSheet1' in book.sheetnames:
    mas1 = book['MasterSheet1']
    book.remove(mas1)
book.create_sheet("MasterSheet1")
row_num = book['MasterSheet'].max_row
column2 = book['MasterSheet'].max_column
mas1 = book['MasterSheet1']
mas1.cell(row=1, column=1).value = row_num
mas1.cell(row=2, column=1).value = column2
book.save("New1.xlsx")
